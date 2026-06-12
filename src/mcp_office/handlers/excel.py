from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import xlsxwriter
import pandas as pd

from mcp_office.base import hex_to_rgb, rgb_to_hex
from mcp_office.handlers.base import DocumentHandler
from mcp_office.spec import SpecManager
from mcp.types import Tool, TextContent


class ExcelHandler(DocumentHandler):
    """Handler for Excel operations"""

    def get_tools(self) -> list[Tool]:
        return [
            Tool(
                name="excel_create",
                description="Create a new Excel workbook",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "path": {
                            "type": "string",
                            "description": "Path to save workbook",
                        },
                        "sheet_name": {"type": "string", "description": "Sheet name"},
                    },
                    "required": ["path"],
                },
            ),
            Tool(
                name="excel_read",
                description="Read data from Excel workbook",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to workbook"},
                        "sheet": {"type": "string", "description": "Sheet name"},
                    },
                    "required": ["path"],
                },
            ),
            Tool(
                name="excel_write_cell",
                description="Write value to a cell",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to workbook"},
                        "sheet": {"type": "string", "description": "Sheet name"},
                        "cell": {"type": "string", "description": "Cell (e.g., A1)"},
                        "value": {"type": "string", "description": "Value"},
                    },
                    "required": ["path", "sheet", "cell", "value"],
                },
            ),
            Tool(
                name="excel_add_row",
                description="Add a row of data",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to workbook"},
                        "sheet": {"type": "string", "description": "Sheet name"},
                        "data": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Row data",
                        },
                    },
                    "required": ["path", "sheet", "data"],
                },
            ),
            Tool(
                name="excel_add_formula",
                description="Add a formula to a cell",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to workbook"},
                        "sheet": {"type": "string", "description": "Sheet name"},
                        "cell": {"type": "string", "description": "Cell"},
                        "formula": {"type": "string", "description": "Formula"},
                    },
                    "required": ["path", "sheet", "cell", "formula"],
                },
            ),
            Tool(
                name="excel_format_cell",
                description="Format a cell",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to workbook"},
                        "sheet": {"type": "string", "description": "Sheet name"},
                        "cell": {"type": "string", "description": "Cell"},
                        "bold": {"type": "boolean", "description": "Bold"},
                        "font_size": {"type": "integer", "description": "Font size"},
                        "font_color": {
                            "type": "string",
                            "description": "Font color hex",
                        },
                        "bg_color": {
                            "type": "string",
                            "description": "Background color",
                        },
                    },
                    "required": ["path", "sheet", "cell"],
                },
            ),
            Tool(
                name="excel_add_table",
                description="Add a table to a sheet",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to workbook"},
                        "sheet": {"type": "string", "description": "Sheet name"},
                        "data": {
                            "type": "array",
                            "items": {"type": "array", "items": {"type": "string"}},
                        },
                        "start_cell": {"type": "string", "description": "Start cell"},
                    },
                    "required": ["path", "sheet", "data", "start_cell"],
                },
            ),
            Tool(
                name="excel_set_column_width",
                description="Set column width",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to workbook"},
                        "sheet": {"type": "string", "description": "Sheet name"},
                        "column": {"type": "string", "description": "Column letter"},
                        "width": {"type": "number", "description": "Width"},
                    },
                    "required": ["path", "sheet", "column", "width"],
                },
            ),
            Tool(
                name="excel_merge_cells",
                description="Merge cells",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to workbook"},
                        "sheet": {"type": "string", "description": "Sheet name"},
                        "start_cell": {"type": "string", "description": "Start cell"},
                        "end_cell": {"type": "string", "description": "End cell"},
                    },
                    "required": ["path", "sheet", "start_cell", "end_cell"],
                },
            ),
            Tool(
                name="excel_list_sheets",
                description="List all sheets",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to workbook"},
                    },
                    "required": ["path"],
                },
            ),
            Tool(
                name="excel_to_pdf",
                description="Convert Excel to PDF",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "input_path": {"type": "string", "description": "Excel path"},
                        "output_path": {"type": "string", "description": "PDF path"},
                    },
                    "required": ["input_path", "output_path"],
                },
            ),
            Tool(
                name="excel_create_xlsxwriter",
                description="Create Excel with xlsxwriter (more formatting)",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "path": {
                            "type": "string",
                            "description": "Path to save workbook",
                        },
                        "sheet_name": {"type": "string", "description": "Sheet name"},
                    },
                    "required": ["path"],
                },
            ),
            Tool(
                name="excel_add_chart",
                description="Add chart to Excel sheet",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to workbook"},
                        "sheet": {"type": "string", "description": "Sheet name"},
                        "chart_type": {
                            "type": "string",
                            "description": "Chart type: column, line, pie, bar, area, scatter, stock",
                        },
                        "data_range": {
                            "type": "string",
                            "description": "Data range (e.g., A1:B5)",
                        },
                        "title": {"type": "string", "description": "Chart title"},
                    },
                    "required": ["path", "chart_type", "data_range"],
                },
            ),
            Tool(
                name="excel_conditional_format",
                description="Add conditional formatting",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to workbook"},
                        "sheet": {"type": "string", "description": "Sheet name"},
                        "range": {"type": "string", "description": "Cell range"},
                        "condition": {
                            "type": "string",
                            "description": "Condition type",
                        },
                        "format": {"type": "string", "description": "Format JSON"},
                    },
                    "required": ["path", "sheet", "range", "condition"],
                },
            ),
            Tool(
                name="excel_pandas_read",
                description="Read Excel with pandas (returns DataFrame)",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to workbook"},
                        "sheet": {"type": "string", "description": "Sheet name"},
                        "header": {"type": "integer", "description": "Header row"},
                    },
                    "required": ["path"],
                },
            ),
            Tool(
                name="excel_pandas_to_excel",
                description="Create Excel from pandas DataFrame",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Output path"},
                        "data": {"type": "string", "description": "JSON data"},
                        "sheet_name": {"type": "string", "description": "Sheet name"},
                    },
                    "required": ["path", "data"],
                },
            ),
        ]

    async def execute(self, tool_name: str, arguments: dict) -> list[TextContent]:
        try:
            if tool_name == "excel_create":
                return self._create(arguments)
            elif tool_name == "excel_read":
                return self._read(arguments)
            elif tool_name == "excel_write_cell":
                return self._write_cell(arguments)
            elif tool_name == "excel_add_row":
                return self._add_row(arguments)
            elif tool_name == "excel_add_formula":
                return self._add_formula(arguments)
            elif tool_name == "excel_format_cell":
                return self._format_cell(arguments)
            elif tool_name == "excel_add_table":
                return self._add_table(arguments)
            elif tool_name == "excel_set_column_width":
                return self._set_column_width(arguments)
            elif tool_name == "excel_merge_cells":
                return self._merge_cells(arguments)
            elif tool_name == "excel_list_sheets":
                return self._list_sheets(arguments)
            elif tool_name == "excel_to_pdf":
                return self._excel_to_pdf(arguments)
            elif tool_name == "excel_create_xlsxwriter":
                return self._create_xlsxwriter(arguments)
            elif tool_name == "excel_add_chart":
                return self._add_chart(arguments)
            elif tool_name == "excel_conditional_format":
                return self._conditional_format(arguments)
            elif tool_name == "excel_pandas_read":
                return self._pandas_read(arguments)
            elif tool_name == "excel_pandas_to_excel":
                return self._pandas_to_excel(arguments)
            return self.error_result(f"Unknown tool: {tool_name}")
        except Exception as e:
            return self.error_result(str(e))

    def _get_sheet(self, wb, sheet_name: str):
        return wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    def _create(self, args: dict) -> list[TextContent]:
        wb = Workbook()
        sheet_name = args.get("sheet_name", "Sheet")
        wb.active.title = sheet_name
        wb.save(args["path"])
        spec = SpecManager("excel", args["path"])
        spec.set_property("sheet_name", sheet_name)
        spec.append({"type": "create", "sheet_name": sheet_name})
        return self.success_result(f"Workbook created: {args['path']}")

    def _read(self, args: dict) -> list[TextContent]:
        import json

        wb = load_workbook(args["path"], data_only=True)
        ws = self._get_sheet(wb, args.get("sheet", ""))
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append([str(cell) if cell is not None else "" for cell in row])
        return [TextContent(type="text", text=json.dumps(data, indent=2))]

    def _write_cell(self, args: dict) -> list[TextContent]:
        wb = load_workbook(args["path"])
        ws = self._get_sheet(wb, args["sheet"])
        ws[args["cell"]] = args["value"]
        wb.save(args["path"])
        SpecManager("excel", args["path"]).append({
            "type": "write_cell",
            "sheet": args["sheet"],
            "cell": args["cell"],
            "value": args["value"],
        })
        return self.success_result(f"Wrote {args['value']} to {args['cell']}")

    def _add_row(self, args: dict) -> list[TextContent]:
        wb = load_workbook(args["path"])
        ws = self._get_sheet(wb, args["sheet"])
        ws.append(args["data"])
        wb.save(args["path"])
        SpecManager("excel", args["path"]).append({
            "type": "add_row",
            "sheet": args["sheet"],
            "data": args["data"],
        })
        return self.success_result(f"Added row with {len(args['data'])} cells")

    def _add_formula(self, args: dict) -> list[TextContent]:
        wb = load_workbook(args["path"])
        ws = self._get_sheet(wb, args["sheet"])
        ws[args["cell"]] = args["formula"]
        wb.save(args["path"])
        SpecManager("excel", args["path"]).append({
            "type": "add_formula",
            "sheet": args["sheet"],
            "cell": args["cell"],
            "formula": args["formula"],
        })
        return self.success_result(f"Added formula {args['formula']}")

    def _format_cell(self, args: dict) -> list[TextContent]:
        wb = load_workbook(args["path"])
        ws = self._get_sheet(wb, args["sheet"])
        cell = ws[args["cell"]]

        if args.get("bold"):
            cell.font = Font(bold=True)
        if size := args.get("font_size"):
            cell.font = Font(size=size)
        if fc := args.get("font_color"):
            rgb = hex_to_rgb(fc)
            cell.font = Font(color=rgb_to_hex(rgb))
        if bg := args.get("bg_color"):
            rgb = hex_to_rgb(bg)
            cell.fill = PatternFill(start_color=rgb_to_hex(rgb), fill_type="solid")

        wb.save(args["path"])
        SpecManager("excel", args["path"]).append({
            "type": "format_cell",
            "sheet": args["sheet"],
            "cell": args["cell"],
            "bold": args.get("bold"),
            "font_size": args.get("font_size"),
            "font_color": args.get("font_color"),
            "bg_color": args.get("bg_color"),
        })
        return self.success_result(f"Formatted cell {args['cell']}")

    def _add_table(self, args: dict) -> list[TextContent]:
        wb = load_workbook(args["path"])
        ws = self._get_sheet(wb, args["sheet"])
        data = args["data"]
        start = ws[args["start_cell"]]

        for r_idx, row in enumerate(data):
            for c_idx, val in enumerate(row):
                cell = ws.cell(row=start.row + r_idx, column=start.column + c_idx)
                cell.value = val

        wb.save(args["path"])
        SpecManager("excel", args["path"]).append({
            "type": "add_table",
            "sheet": args["sheet"],
            "data": data,
            "start_cell": args["start_cell"],
        })
        return self.success_result(f"Added table with {len(data)} rows")

    def _set_column_width(self, args: dict) -> list[TextContent]:
        wb = load_workbook(args["path"])
        ws = self._get_sheet(wb, args["sheet"])
        ws.column_dimensions[args["column"]].width = args["width"]
        wb.save(args["path"])
        SpecManager("excel", args["path"]).append({
            "type": "set_column_width",
            "sheet": args["sheet"],
            "column": args["column"],
            "width": args["width"],
        })
        return self.success_result(f"Set column {args['column']} width")

    def _merge_cells(self, args: dict) -> list[TextContent]:
        wb = load_workbook(args["path"])
        ws = self._get_sheet(wb, args["sheet"])
        ws.merge_cells(f"{args['start_cell']}:{args['end_cell']}")
        wb.save(args["path"])
        SpecManager("excel", args["path"]).append({
            "type": "merge_cells",
            "sheet": args["sheet"],
            "start_cell": args["start_cell"],
            "end_cell": args["end_cell"],
        })
        return self.success_result(f"Merged {args['start_cell']}:{args['end_cell']}")

    def _list_sheets(self, args: dict) -> list[TextContent]:
        import json

        wb = load_workbook(args["path"])
        return [TextContent(type="text", text=json.dumps(wb.sheetnames, indent=2))]

    def _excel_to_pdf(self, args: dict) -> list[TextContent]:
        SpecManager("excel", args["input_path"]).add_post_processing(
            "convert_to_pdf", output=args.get("output_path")
        )
        return self.success_result(
            "Excel to PDF requires LibreOffice. Install and use: soffice --headless --convert-to pdf input.xlsx"
        )

    def _create_xlsxwriter(self, args: dict) -> list[TextContent]:
        workbook = xlsxwriter.Workbook(args["path"])
        workbook.add_worksheet(args.get("sheet_name", "Sheet"))
        workbook.close()
        spec = SpecManager("excel", args["path"])
        spec.set_property("sheet_name", args.get("sheet_name", "Sheet"))
        spec.append({"type": "create_xlsxwriter", "sheet_name": args.get("sheet_name", "Sheet")})
        return self.success_result(f"Workbook created: {args['path']}")

    def _add_chart(self, args: dict) -> list[TextContent]:
        import xlsxwriter

        workbook = xlsxwriter.Workbook(args["path"], {"in_memory": True})
        ws = workbook.add_worksheet(args.get("sheet", "Sheet"))

        chart_types = {
            "column": "column",
            "line": "line",
            "pie": "pie",
            "bar": "bar",
            "area": "area",
            "scatter": "scatter",
            "stock": "stock",
        }
        chart_type = chart_types.get(args.get("chart_type", "column"), "column")

        chart = workbook.add_chart({"type": chart_type})
        chart.add_series({"values": args.get("data_range", "Sheet!$A$1:$B$5")})

        if title := args.get("title"):
            chart.set_title({"name": title})

        ws.insert_chart("E2", chart)
        workbook.close()
        SpecManager("excel", args["path"]).append({
            "type": "add_chart",
            "sheet": args.get("sheet", "Sheet"),
            "chart_type": args.get("chart_type"),
            "data_range": args.get("data_range"),
            "title": args.get("title"),
        })
        return self.success_result(f"Chart added: {args.get('chart_type')}")

    def _conditional_format(self, args: dict) -> list[TextContent]:
        wb = load_workbook(args["path"])
        ws = self._get_sheet(wb, args["sheet"])

        format_map = {
            "greater": "greater_than",
            "less": "less_than",
            "between": "between",
            "equal": "equal",
            "contains": "text_contains",
        }
        condition = format_map.get(args.get("condition", "greater"), "greater_than")

        xlsx_format = wb.add_format({"bg_color": "#FFC7CE", "font_color": "#9C0006"})
        ws.conditional_format(
            args["range"], {"type": condition, "value": 0, "format": xlsx_format}
        )

        wb.save(args["path"])
        SpecManager("excel", args["path"]).append({
            "type": "conditional_format",
            "sheet": args["sheet"],
            "range": args["range"],
            "condition": args.get("condition"),
        })
        return self.success_result("Conditional formatting added")

    def _pandas_read(self, args: dict) -> list[TextContent]:
        df = pd.read_excel(
            args["path"], sheet_name=args.get("sheet"), header=args.get("header")
        )
        return [TextContent(type="text", text=df.to_json(orient="records", indent=2))]

    def _pandas_to_excel(self, args: dict) -> list[TextContent]:
        import json

        data = json.loads(args["data"])
        df = pd.DataFrame(data)
        df.to_excel(
            args["path"], sheet_name=args.get("sheet_name", "Sheet"), index=False
        )
        spec = SpecManager("excel", args["path"])
        spec.set_property("sheet_name", args.get("sheet_name", "Sheet"))
        spec.append({"type": "pandas_to_excel", "data_preview": str(data[:3]) if data else "[]"})
        return self.success_result(f"Created Excel from data: {args['path']}")
